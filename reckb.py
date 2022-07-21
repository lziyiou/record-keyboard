from pynput import keyboard
from openpyxl import load_workbook,Workbook
import os


filepath = os.path.dirname(__file__) + '/key_count.xlsx'
key_count_dic = {}


def on_press(key):
    # press F12 to exit
    if key == keyboard.Key.f12:
        # create a new file or open the file
        try:
            workbook = load_workbook(filepath)
        except:
            workbook = Workbook()
        ws = workbook.active

        # write to the file
        for k,v in key_count_dic.items():
            i = 1
            while True:
                cellA = ws.cell(row=i, column=1)
                if cellA.value == k:
                    cellB = ws.cell(row=i, column=2)
                    cellB.value += v
                    break
                elif cellA.value == None:
                    cellB = ws.cell(row=i, column=2)
                    cellB.value = v
                    cellA.value = str(k)
                    break
                i += 1

        # save the file
        workbook.save('key_count.xlsx')

        return False
    
    else:
        if type(key) == keyboard._xorg.KeyCode:
            key = key.char
        key_count_dic[key] = key_count_dic.get(key, 0)+1

def on_release(key):
    pass

with keyboard.Listener(
        on_press=on_press,
        on_release=on_release
        ) as listener:
    listener.join()

